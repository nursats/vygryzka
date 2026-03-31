import os
import json
import time
import datetime
import requests
import openpyxl

EXCEL_FILE = "20260304 KOMPRA (2).xlsx"
SHEET_NAME = "Phase 1"
API_TOKEN = "G1BCW0TvPlFFf7jM7wmLoi"
API_URL = "https://kompra.kz/api/v2/tax-details"
JSON_DIR = "json_responses"
OUTPUT_EXCEL = "kompra_tax_details.xlsx"
MAX_RETRIES = 2
RETRY_DELAY = 2  # seconds

# Track progress
PROGRESS_FILE = "progress.json"


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r") as f:
            return json.load(f)
    return {"completed": []}


def save_progress(progress):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f)


def read_bins():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb[SHEET_NAME]
    bins = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            bins.append(str(val).strip())
    wb.close()
    return bins


def fetch_tax_details(identifier):
    """Fetch tax details for a given BIN. Retries if response has empty payment list with null names."""
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.get(
                API_URL,
                params={"identifier": identifier, "api-token": API_TOKEN},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()

            # status:null means data is still loading -> retry
            # status:false means BIN has no tax data -> save as-is
            top_status = data.get("status")
            content = data.get("content", {})
            answer = content.get("answer", {})
            payments = answer.get("payment", [])

            if len(payments) == 0 and top_status is None:
                if attempt < MAX_RETRIES - 1:
                    print(f"    Empty response (status=null) for {identifier}, retrying in {RETRY_DELAY}s (attempt {attempt + 1}/{MAX_RETRIES})...")
                    time.sleep(RETRY_DELAY)
                    continue
                else:
                    print(f"    Max retries reached for {identifier}, skipping")
                    return None

            if len(payments) == 0 and top_status is False:
                print(f"    No tax data for {identifier} (status=false), saving")

            return data

        except requests.exceptions.RequestException as e:
            if attempt < MAX_RETRIES - 1:
                print(f"    Request error for {identifier}: {e}, retrying in {RETRY_DELAY}s...")
                time.sleep(RETRY_DELAY)
            else:
                print(f"    Failed after {MAX_RETRIES} attempts for {identifier}: {e}")
                return None
    return None


def timestamp_to_date(ts):
    """Convert millisecond timestamp to YYYY-MM-DD string."""
    if ts is None:
        return ""
    try:
        return datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
    except (ValueError, OSError):
        return ""


def json_to_rows(identifier, data):
    """Convert API JSON response to rows for Excel."""
    rows = []
    if data is None:
        return rows

    content = data.get("content", {})
    answer = content.get("answer", {})
    payments = answer.get("payment", [])

    for p in payments:
        rows.append([
            identifier,
            p.get("summa", ""),
            timestamp_to_date(p.get("receiptDate")),
            timestamp_to_date(p.get("writeOffDate")),
            p.get("year", ""),
            p.get("payType", ""),
            p.get("taxOrgCode", ""),
            p.get("nameTaxRu", ""),
            p.get("nameTaxKz", ""),
            p.get("kbk", ""),
            p.get("kbkNameRu", ""),
            p.get("kbkNameKz", ""),
        ])
    return rows


def build_excel_from_jsons():
    """Read all saved JSONs and build the final Excel file."""
    print("\nBuilding final Excel file...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Details"

    headers = [
        "identifier", "summa", "receipt_date", "write_off_date", "year",
        "pay_type", "tax_org_code", "name_ru", "name_kz",
        "kbk_code", "name_ru", "name_kz"
    ]
    ws.append(headers)

    json_files = sorted(f for f in os.listdir(JSON_DIR) if f.endswith(".json"))
    total_rows = 0
    for jf in json_files:
        identifier = jf.replace(".json", "")
        with open(os.path.join(JSON_DIR, jf), "r") as f:
            data = json.load(f)
        rows = json_to_rows(identifier, data)
        for row in rows:
            ws.append(row)
            total_rows += 1

    wb.save(OUTPUT_EXCEL)
    print(f"Saved {OUTPUT_EXCEL} with {total_rows} rows from {len(json_files)} BINs")


def main():
    os.makedirs(JSON_DIR, exist_ok=True)

    # Load progress to resume from where we left off
    progress = load_progress()
    completed_set = set(progress["completed"])

    bins = read_bins()
    print(f"Total BINs to process: {len(bins)}")
    print(f"Already completed: {len(completed_set)}")

    remaining = [b for b in bins if b not in completed_set]
    print(f"Remaining: {len(remaining)}")

    for i, bin_id in enumerate(remaining):
        print(f"[{len(completed_set) + i + 1}/{len(bins)}] Fetching {bin_id}...")
        data = fetch_tax_details(bin_id)

        if data is not None:
            json_path = os.path.join(JSON_DIR, f"{bin_id}.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            # Only mark as completed if we got real data
            progress["completed"].append(bin_id)
            completed_set.add(bin_id)
            save_progress(progress)
        else:
            print(f"    Skipped {bin_id} (no data), will retry next run")

        # Small delay to be respectful to the API
        time.sleep(0.5)

    print(f"\nAll {len(bins)} BINs processed!")

    # Build final Excel
    build_excel_from_jsons()


if __name__ == "__main__":
    main()
