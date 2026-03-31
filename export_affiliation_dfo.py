import argparse
import os
import re
from typing import Dict, Iterable, List, Sequence

import openpyxl
from sqlalchemy import bindparam, create_engine, text
from sqlalchemy.engine import Engine

DEFAULT_INPUT_EXCEL = "20260304 KOMPRA (2).xlsx"
DEFAULT_INPUT_DIR = "new"
DEFAULT_OUTPUT_DIR = "affiliation_dfo_excels"
DEFAULT_SHEET_NAME = "Phase 1"
DEFAULT_START_ROW = 3
DEFAULT_COLUMN = 1
DEFAULT_CHUNK_SIZE = 1000

# table_name -> BIN filter column
TABLE_FILTERS: Dict[str, str] = {
    "affiliation": "src_identifier",
    "executives": "src_identifier",
    "board_members": "src_identifier",
    "accountant": "src_identifier",
    "dfo_status": "identifier",
    "dfo_status_shadow": "identifier",
}


def normalize_bin(value) -> str:
    if value is None:
        return ""

    if isinstance(value, int):
        raw = str(value)
    elif isinstance(value, float):
        raw = str(int(value)) if value.is_integer() else str(value)
    else:
        raw = str(value).strip()

    raw = raw.replace("\u00a0", "").replace(" ", "")
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]

    if raw.isdigit() and len(raw) <= 12:
        return raw.zfill(12)

    return raw


def read_bins_from_excel(
    file_path: str,
    sheet_name: str,
    start_row: int,
    column: int,
) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"Sheet '{sheet_name}' not found. Using '{ws.title}' instead.")

    bins: List[str] = []
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row,
        min_col=column,
        max_col=column,
        values_only=True,
    ):
        value = normalize_bin(row[0])
        if value:
            bins.append(value)

    wb.close()

    # Keep source order but drop duplicates.
    seen = set()
    unique_bins: List[str] = []
    for bin_value in bins:
        if bin_value not in seen:
            unique_bins.append(bin_value)
            seen.add(bin_value)

    return unique_bins


def read_bins_from_dir(
    dir_path: str,
    sheet_name: str,
    start_row: int,
    column: int,
) -> List[str]:
    excel_paths = sorted(
        os.path.join(dir_path, name)
        for name in os.listdir(dir_path)
        if name.lower().endswith(".xlsx") and not name.startswith("~$")
    )
    if not excel_paths:
        raise FileNotFoundError(f"No .xlsx files found in directory: {dir_path}")

    bins: List[str] = []
    seen = set()

    for excel_path in excel_paths:
        file_bins = read_bins_from_excel(
            file_path=excel_path,
            sheet_name=sheet_name,
            start_row=start_row,
            column=column,
        )
        print(f"{os.path.basename(excel_path)}: {len(file_bins)} unique BINs")
        for bin_value in file_bins:
            if bin_value not in seen:
                bins.append(bin_value)
                seen.add(bin_value)

    return bins


def get_db_settings() -> dict:
    return {
        "name": os.getenv("DB_NAME", "kompra"),
        "user": os.getenv("DB_USER", "python"),
        "password": os.getenv("DB_PASS", "python123"),
        "host": os.getenv("DB_HOST", "10.0.0.1"),
        "port": os.getenv("DB_PORT", "5432"),
        "schema": os.getenv("DB_SCHEMA", "public"),
    }


def create_db_engine(settings: dict) -> Engine:
    url = "postgresql://{}:{}@{}:{}/{}".format(
        settings["user"],
        settings["password"],
        settings["host"],
        settings["port"],
        settings["name"],
    )
    return create_engine(url, client_encoding="utf8")


def validate_sql_identifier(name: str, kind: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name):
        raise ValueError(f"Invalid {kind}: {name!r}")
    return name


def chunked(values: Sequence[str], chunk_size: int) -> Iterable[Sequence[str]]:
    for i in range(0, len(values), chunk_size):
        yield values[i : i + chunk_size]


def fetch_table_columns(engine: Engine, schema: str, table_name: str) -> List[str]:
    sql = text(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = :schema
          AND table_name = :table
        ORDER BY ordinal_position
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(sql, {"schema": schema, "table": table_name}).all()
    return [row[0] for row in rows]


def fetch_table_rows(
    engine: Engine,
    schema: str,
    table_name: str,
    filter_column: str,
    bins: Sequence[str],
    chunk_size: int,
) -> List[dict]:
    if not bins:
        return []

    stmt = text(
        f"""
        SELECT *
        FROM {schema}.{table_name}
        WHERE {filter_column} IN :bins
        ORDER BY id
        """
    ).bindparams(bindparam("bins", expanding=True))

    result_rows: List[dict] = []
    with engine.connect() as conn:
        for bin_chunk in chunked(bins, chunk_size):
            rows = conn.execute(stmt, {"bins": list(bin_chunk)}).mappings().all()
            result_rows.extend(dict(row) for row in rows)

    return result_rows


def _strip_tz(value):
    """Remove timezone info from datetime/time so openpyxl can write it."""
    import datetime
    if isinstance(value, (datetime.datetime, datetime.time)) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


EXCLUDE_COLUMNS = {"id", "created", "last_updated"}


def write_table_to_excel(output_path: str, columns: Sequence[str], rows: Sequence[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"

    cols = [c for c in columns if c not in EXCLUDE_COLUMNS]
    ws.append(cols)
    for row in rows:
        ws.append([_strip_tz(row.get(col)) for col in cols])

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export Affiliation/DFO related tables by BIN list from Excel. "
            "Each table is saved to a separate Excel file."
        )
    )
    parser.add_argument("--input", default=DEFAULT_INPUT_EXCEL, help="Path to input Excel with BINs")
    parser.add_argument(
        "--input-dir",
        default=None,
        help="Directory with input Excels containing BINs in the selected column",
    )
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for output Excels (one file per table)",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Sheet name with BINs")
    parser.add_argument("--start-row", type=int, default=DEFAULT_START_ROW, help="Data start row")
    parser.add_argument("--column", type=int, default=DEFAULT_COLUMN, help="BIN column number (1-based)")
    parser.add_argument("--schema", default=None, help="DB schema (default: DB_SCHEMA env or public)")
    parser.add_argument("--chunk-size", type=int, default=DEFAULT_CHUNK_SIZE, help="BIN query chunk size")
    parser.add_argument(
        "--tables",
        default=",".join(TABLE_FILTERS.keys()),
        help="Comma-separated tables to export",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.input_dir:
        if not os.path.isdir(args.input_dir):
            raise FileNotFoundError(f"Input directory not found: {args.input_dir}")
        bins = read_bins_from_dir(
            dir_path=args.input_dir,
            sheet_name=args.sheet,
            start_row=args.start_row,
            column=args.column,
        )
        print(f"Read {len(bins)} unique BINs from directory")
    else:
        if not os.path.exists(args.input):
            raise FileNotFoundError(f"Input file not found: {args.input}")

        bins = read_bins_from_excel(
            file_path=args.input,
            sheet_name=args.sheet,
            start_row=args.start_row,
            column=args.column,
        )
        print(f"Read {len(bins)} unique BINs from Excel")

    settings = get_db_settings()
    if args.schema:
        settings["schema"] = args.schema

    schema = validate_sql_identifier(settings["schema"], "schema")

    requested_tables = [t.strip() for t in args.tables.split(",") if t.strip()]
    if not requested_tables:
        raise ValueError("No tables selected for export")

    for table_name in requested_tables:
        if table_name not in TABLE_FILTERS:
            supported = ", ".join(TABLE_FILTERS.keys())
            raise ValueError(f"Unsupported table '{table_name}'. Supported: {supported}")

    engine = create_db_engine(settings)
    os.makedirs(args.output_dir, exist_ok=True)

    files_created = 0
    for table_name in requested_tables:
        validate_sql_identifier(table_name, "table")
        filter_column = validate_sql_identifier(TABLE_FILTERS[table_name], "column")

        columns = fetch_table_columns(engine, schema, table_name)
        if not columns:
            print(f"Skip {table_name}: table not found in schema '{schema}'")
            continue

        rows = fetch_table_rows(
            engine=engine,
            schema=schema,
            table_name=table_name,
            filter_column=filter_column,
            bins=bins,
            chunk_size=max(1, args.chunk_size),
        )

        if not rows:
            print(f"{table_name}: 0 rows (no file)")
            continue

        # Put filter (BIN) column first
        if filter_column in columns:
            columns = [filter_column] + [c for c in columns if c != filter_column]

        output_name = f"{table_name}.xlsx"
        if args.input_dir:
            output_name = f"{table_name}_new.xlsx"

        output_path = os.path.join(args.output_dir, output_name)
        write_table_to_excel(output_path, columns, rows)
        files_created += 1
        print(f"{table_name}: {len(rows)} rows -> {output_path}")

    if files_created == 0:
        raise RuntimeError("No data found for selected BINs in selected tables.")

    print(f"Created {files_created} files in: {args.output_dir}")


if __name__ == "__main__":
    main()
