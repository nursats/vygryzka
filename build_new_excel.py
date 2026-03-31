import datetime
import json
import os

import openpyxl

JSON_DIR = "json_responses_new"
OUTPUT_TAX = "new_1_Налоговые_отчисления.xlsx"
OUTPUT_FOT = "new_2_Оценочный_ФОТ.xlsx"
OUTPUT_FINES = "new_3_Штрафы_и_пени.xlsx"
MAX_EXCEL_ROWS = 1_048_576
MAX_DATA_ROWS_PER_SHEET = MAX_EXCEL_ROWS - 1


def timestamp_to_date(ts):
    if ts is None:
        return ""
    try:
        return datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
    except (ValueError, OSError):
        return ""


def load_all_payments():
    payments = []
    json_files = sorted(f for f in os.listdir(JSON_DIR) if f.endswith(".json"))
    print(f"Found {len(json_files)} JSON files")

    for filename in json_files:
        identifier = filename[:-5]
        path = os.path.join(JSON_DIR, filename)
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"Skipping {filename}: {e}")
            continue

        for payment in data.get("content", {}).get("answer", {}).get("payment", []):
            payments.append(
                {
                    "identifier": identifier,
                    "summa": payment.get("summa", 0) or 0,
                    "receipt_date": timestamp_to_date(payment.get("receiptDate")),
                    "write_off_date": timestamp_to_date(payment.get("writeOffDate")),
                    "year": payment.get("year", ""),
                    "pay_type": payment.get("payType", ""),
                    "tax_org_code": payment.get("taxOrgCode", ""),
                    "name_ru": payment.get("nameTaxRu", ""),
                    "name_kz": payment.get("nameTaxKz", ""),
                    "kbk_code": payment.get("kbk", ""),
                    "kbk_name_ru": payment.get("kbkNameRu", ""),
                    "kbk_name_kz": payment.get("kbkNameKz", ""),
                }
            )

    print(f"Total payment rows: {len(payments)}")
    return payments


def append_rows_with_sheet_split(workbook, sheet_title, headers, rows):
    sheet_index = 1
    ws = workbook.active
    ws.title = sheet_title
    ws.append(headers)
    written_on_sheet = 0

    for row in rows:
        if written_on_sheet >= MAX_DATA_ROWS_PER_SHEET:
            sheet_index += 1
            ws = workbook.create_sheet(f"{sheet_title} {sheet_index}")
            ws.append(headers)
            written_on_sheet = 0

        ws.append(row)
        written_on_sheet += 1


def build_tax_excel(payments):
    wb = openpyxl.Workbook()
    headers = [
        "identifier",
        "summa",
        "receipt_date",
        "write_off_date",
        "year",
        "pay_type",
        "tax_org_code",
        "name_ru",
        "name_kz",
        "kbk_code",
        "kbk_name_ru",
        "kbk_name_kz",
    ]
    rows = []
    for payment in payments:
        rows.append(
            [
                payment["identifier"],
                payment["summa"],
                payment["receipt_date"],
                payment["write_off_date"],
                payment["year"],
                payment["pay_type"],
                payment["tax_org_code"],
                payment["name_ru"],
                payment["name_kz"],
                payment["kbk_code"],
                payment["kbk_name_ru"],
                payment["kbk_name_kz"],
            ]
        )
    append_rows_with_sheet_split(wb, "Налоговые отчисления", headers, rows)
    wb.save(OUTPUT_TAX)
    print(f"Saved {OUTPUT_TAX}")


def build_fot_excel(payments):
    wb = openpyxl.Workbook()
    headers = ["identifier", "summa", "write_off_date", "year"]
    rows = []
    for payment in payments:
        if payment["kbk_code"] == "901101" and payment["pay_type"] == 1:
            rows.append(
                [
                    payment["identifier"],
                    round(payment["summa"] * 10, 2),
                    payment["write_off_date"],
                    payment["year"],
                ]
            )
    append_rows_with_sheet_split(wb, "Оценочный ФОТ", headers, rows)
    wb.save(OUTPUT_FOT)
    print(f"Saved {OUTPUT_FOT}")


def build_fines_excel(payments):
    wb = openpyxl.Workbook()
    headers = [
        "identifier",
        "summa",
        "receipt_date",
        "write_off_date",
        "year",
        "pay_type",
        "tax_org_code",
        "name_ru",
        "name_kz",
        "kbk_code",
        "kbk_name_ru",
        "kbk_name_kz",
    ]
    rows = []
    for payment in payments:
        if payment["pay_type"] in (2, 3):
            rows.append(
                [
                    payment["identifier"],
                    payment["summa"],
                    payment["receipt_date"],
                    payment["write_off_date"],
                    payment["year"],
                    payment["pay_type"],
                    payment["tax_org_code"],
                    payment["name_ru"],
                    payment["name_kz"],
                    payment["kbk_code"],
                    payment["kbk_name_ru"],
                    payment["kbk_name_kz"],
                ]
            )
    append_rows_with_sheet_split(wb, "Штрафы и пени", headers, rows)
    wb.save(OUTPUT_FINES)
    print(f"Saved {OUTPUT_FINES}")


def main():
    payments = load_all_payments()
    build_tax_excel(payments)
    build_fot_excel(payments)
    build_fines_excel(payments)


if __name__ == "__main__":
    main()
