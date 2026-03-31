import datetime
import json
import os
import re
import sys
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

import openpyxl

BIN_LIST_PATH = "all_bins.txt"
DEFAULT_JSON_DIR = "all_bins_json"
OUTPUT_TAX = "all_bins_1_Налоговые_отчисления.xlsx"
OUTPUT_FOT = "all_bins_2_Оценочный_ФОТ.xlsx"
OUTPUT_FINES = "all_bins_3_Штрафы_и_пени.xlsx"
MAX_EXCEL_ROWS = 1_048_576
MAX_DATA_ROWS_PER_SHEET = MAX_EXCEL_ROWS - 1


def timestamp_to_date(ts):
    if ts is None:
        return ""
    try:
        return datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
    except (ValueError, OSError, TypeError):
        return ""


def load_bins(path: str) -> List[str]:
    text = Path(path).read_text(encoding="utf-8")
    bins = re.findall(r"\d+", text)
    # Preserve source order but drop duplicates.
    return list(dict.fromkeys(bins))


def payment_to_row(identifier: str, payment: dict) -> Tuple:
    return (
        identifier,
        payment.get("summa", 0) or 0,
        timestamp_to_date(payment.get("receiptDate")),
        timestamp_to_date(payment.get("writeOffDate")),
        payment.get("year", ""),
        payment.get("payType", ""),
        payment.get("taxOrgCode", ""),
        payment.get("nameTaxRu", ""),
        payment.get("nameTaxKz", ""),
        payment.get("kbk", ""),
        payment.get("kbkNameRu", ""),
        payment.get("kbkNameKz", ""),
    )


def load_payments_for_bins(bins: Sequence[str], json_dir: str):
    payments = []
    found_bins = set()
    seen_rows = set()

    for identifier in bins:
        path = os.path.join(json_dir, f"{identifier}.json")
        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"Skipping {path}: {e}")
            continue

        found_bins.add(identifier)
        for payment in data.get("content", {}).get("answer", {}).get("payment", []):
            row = payment_to_row(identifier, payment)
            if row in seen_rows:
                continue
            seen_rows.add(row)
            payments.append(
                {
                    "identifier": row[0],
                    "summa": row[1],
                    "receipt_date": row[2],
                    "write_off_date": row[3],
                    "year": row[4],
                    "pay_type": row[5],
                    "tax_org_code": row[6],
                    "name_ru": row[7],
                    "name_kz": row[8],
                    "kbk_code": row[9],
                    "kbk_name_ru": row[10],
                    "kbk_name_kz": row[11],
                }
            )

    return payments, found_bins


def append_rows_with_sheet_split(workbook, sheet_title: str, headers: Sequence[str], rows: Iterable[Sequence]):
    sheet_index = 1
    ws = workbook.active
    ws.title = sheet_title
    ws.append(list(headers))
    written_on_sheet = 0

    for row in rows:
        if written_on_sheet >= MAX_DATA_ROWS_PER_SHEET:
            sheet_index += 1
            ws = workbook.create_sheet(f"{sheet_title} {sheet_index}")
            ws.append(list(headers))
            written_on_sheet = 0

        ws.append(list(row))
        written_on_sheet += 1


def build_tax_excel(payments):
    headers = [
        "identifier",
        "summa",
        "receipt_date",
        "write_off_date",
        "year",
        "pay_type",
        "tax_org_code",
        "name_ru",
        "name_kz",
        "kbk_code",
        "kbk_name_ru",
        "kbk_name_kz",
    ]
    rows = [
        [
            payment["identifier"],
            payment["summa"],
            payment["receipt_date"],
            payment["write_off_date"],
            payment["year"],
            payment["pay_type"],
            payment["tax_org_code"],
            payment["name_ru"],
            payment["name_kz"],
            payment["kbk_code"],
            payment["kbk_name_ru"],
            payment["kbk_name_kz"],
        ]
        for payment in payments
    ]

    wb = openpyxl.Workbook()
    append_rows_with_sheet_split(wb, "Налоговые отчисления", headers, rows)
    wb.save(OUTPUT_TAX)
    print(f"Saved {OUTPUT_TAX}: {len(rows)} rows")


def build_fot_excel(payments):
    headers = ["identifier", "summa", "write_off_date", "year"]
    rows = []
    for payment in payments:
        if payment["kbk_code"] == "901101" and payment["pay_type"] == 1:
            rows.append(
                [
                    payment["identifier"],
                    round(payment["summa"] * 10, 2),
                    payment["write_off_date"],
                    payment["year"],
                ]
            )

    wb = openpyxl.Workbook()
    append_rows_with_sheet_split(wb, "Оценочный ФОТ", headers, rows)
    wb.save(OUTPUT_FOT)
    print(f"Saved {OUTPUT_FOT}: {len(rows)} rows")


def build_fines_excel(payments):
    headers = [
        "identifier",
        "summa",
        "receipt_date",
        "write_off_date",
        "year",
        "pay_type",
        "tax_org_code",
        "name_ru",
        "name_kz",
        "kbk_code",
        "kbk_name_ru",
        "kbk_name_kz",
    ]
    rows = []
    for payment in payments:
        if payment["pay_type"] in (2, 3):
            rows.append(
                [
                    payment["identifier"],
                    payment["summa"],
                    payment["receipt_date"],
                    payment["write_off_date"],
                    payment["year"],
                    payment["pay_type"],
                    payment["tax_org_code"],
                    payment["name_ru"],
                    payment["name_kz"],
                    payment["kbk_code"],
                    payment["kbk_name_ru"],
                    payment["kbk_name_kz"],
                ]
            )

    wb = openpyxl.Workbook()
    append_rows_with_sheet_split(wb, "Штрафы и пени", headers, rows)
    wb.save(OUTPUT_FINES)
    print(f"Saved {OUTPUT_FINES}: {len(rows)} rows")


def main():
    json_dir = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_JSON_DIR
    if not os.path.isdir(json_dir):
        raise FileNotFoundError(
            f"JSON folder not found: {json_dir}. "
            f"Create it or pass a folder path: python3 build_bins_exports.py <json_folder>"
        )

    bins = load_bins(BIN_LIST_PATH)
    payments, found_bins = load_payments_for_bins(bins, json_dir)
    missing_count = len(bins) - len(found_bins)

    print(f"Loaded BINs from {BIN_LIST_PATH}: {len(bins)}")
    print(f"Using JSON folder: {json_dir}")
    print(f"Found JSON data for: {len(found_bins)}")
    print(f"BINs without local JSON data: {missing_count}")
    print(f"Collected unique payment rows: {len(payments)}")

    build_tax_excel(payments)
    build_fot_excel(payments)
    build_fines_excel(payments)


if __name__ == "__main__":
    main()
