import datetime
import json
import os

import openpyxl

JSON_DIR = "all_bins_json"
OUTPUT_TAX = "1_Налоговые_отчисления.xlsx"
OUTPUT_FOT = "2_Оценочный_ФОТ.xlsx"
OUTPUT_FINES = "3_Штрафы_и_пени.xlsx"
MAX_EXCEL_ROWS = 1_048_576
MAX_DATA_ROWS_PER_SHEET = MAX_EXCEL_ROWS - 1


def timestamp_to_date(ts):
    if ts is None:
        return ""
    try:
        return datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
    except (ValueError, OSError, TypeError):
        return ""


def iter_json_files():
    for filename in sorted(f for f in os.listdir(JSON_DIR) if f.endswith(".json")):
        yield filename


def iter_payments():
    json_files = list(iter_json_files())
    print(f"Found {len(json_files)} JSON files")

    total_payments = 0
    for filename in json_files:
        identifier = filename[:-5]
        path = os.path.join(JSON_DIR, filename)

        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"Skipping {filename}: {e}")
            continue

        payments = data.get("content", {}).get("answer", {}).get("payment", [])
        for payment in payments:
            total_payments += 1
            yield {
                "identifier": identifier,
                "summa": payment.get("summa", 0) or 0,
                "receipt_date": timestamp_to_date(payment.get("receiptDate")),
                "write_off_date": timestamp_to_date(payment.get("writeOffDate")),
                "year": payment.get("year", ""),
                "pay_type": payment.get("payType", ""),
                "tax_org_code": payment.get("taxOrgCode", ""),
                "name_ru": payment.get("nameTaxRu", ""),
                "name_kz": payment.get("nameTaxKz", ""),
                "kbk_code": payment.get("kbk", ""),
                "kbk_name_ru": payment.get("kbkNameRu", ""),
                "kbk_name_kz": payment.get("kbkNameKz", ""),
            }

    print(f"Total payment records scanned: {total_payments}")


def create_sheet(workbook, sheet_title, headers, sheet_index):
    name = sheet_title if sheet_index == 1 else f"{sheet_title} {sheet_index}"
    ws = workbook.create_sheet(title=name)
    ws.append(headers)
    return ws


def stream_to_workbook(output_path, sheet_title, headers, row_iterable):
    wb = openpyxl.Workbook(write_only=True)
    sheet_index = 1
    ws = create_sheet(wb, sheet_title, headers, sheet_index)
    written_on_sheet = 0
    total_written = 0

    for row in row_iterable:
        if written_on_sheet >= MAX_DATA_ROWS_PER_SHEET:
            sheet_index += 1
            ws = create_sheet(wb, sheet_title, headers, sheet_index)
            written_on_sheet = 0

        ws.append(row)
        written_on_sheet += 1
        total_written += 1

    wb.save(output_path)
    print(f"Saved {output_path}: {total_written} rows")


def tax_rows():
    for payment in iter_payments():
        yield [
            payment["identifier"],
            payment["summa"],
            payment["receipt_date"],
            payment["write_off_date"],
            payment["year"],
            payment["pay_type"],
            payment["tax_org_code"],
            payment["name_ru"],
            payment["name_kz"],
            payment["kbk_code"],
            payment["kbk_name_ru"],
            payment["kbk_name_kz"],
        ]


def fot_rows():
    for payment in iter_payments():
        if payment["kbk_code"] == "901101" and payment["pay_type"] == 1:
            yield [
                payment["identifier"],
                round(payment["summa"] * 10, 2),
                payment["write_off_date"],
                payment["year"],
            ]


def fines_rows():
    for payment in iter_payments():
        if payment["pay_type"] in (2, 3):
            yield [
                payment["identifier"],
                payment["summa"],
                payment["receipt_date"],
                payment["write_off_date"],
                payment["year"],
                payment["pay_type"],
                payment["tax_org_code"],
                payment["name_ru"],
                payment["name_kz"],
                payment["kbk_code"],
                payment["kbk_name_ru"],
                payment["kbk_name_kz"],
            ]


def main():
    stream_to_workbook(
        OUTPUT_TAX,
        "Налоговые отчисления",
        [
            "identifier",
            "summa",
            "receipt_date",
            "write_off_date",
            "year",
            "pay_type",
            "tax_org_code",
            "name_ru",
            "name_kz",
            "kbk_code",
            "kbk_name_ru",
            "kbk_name_kz",
        ],
        tax_rows(),
    )

    stream_to_workbook(
        OUTPUT_FOT,
        "Оценочный ФОТ",
        ["identifier", "summa", "write_off_date", "year"],
        fot_rows(),
    )

    stream_to_workbook(
        OUTPUT_FINES,
        "Штрафы и пени",
        [
            "identifier",
            "summa",
            "receipt_date",
            "write_off_date",
            "year",
            "pay_type",
            "tax_org_code",
            "name_ru",
            "name_kz",
            "kbk_code",
            "kbk_name_ru",
            "kbk_name_kz",
        ],
        fines_rows(),
    )

    print("\nDone! All 3 Excel files created.")


if __name__ == "__main__":
    main()
