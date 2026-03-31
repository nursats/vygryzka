import os
from typing import List, Sequence

import openpyxl

TAX_SOURCE_FILES = {
    "1_Налоговые_отчисления.xlsx": "Налоговые отчисления",
    "2_Оценочный_ФОТ.xlsx": "Оценочный ФОТ",
    "3_Штрафы_и_пени.xlsx": "Штрафы и пени",
}
NEW_TAX_SOURCE_FILES = {
    "1_Налоговые_отчисления.xlsx": "new_1_Налоговые_отчисления.xlsx",
    "2_Оценочный_ФОТ.xlsx": "new_2_Оценочный_ФОТ.xlsx",
    "3_Штрафы_и_пени.xlsx": "new_3_Штрафы_и_пени.xlsx",
}
FINANCIAL_FILES = ["bin_financial.xlsx", "bin_financial_new.xlsx"]
OUTPUT_DIR = "combined_excels"
MAX_EXCEL_ROWS = 1_048_576
MAX_DATA_ROWS_PER_SHEET = MAX_EXCEL_ROWS - 1

AFFILIATION_OUTPUTS = {
    "affiliation": "affiliation.xlsx",
    "executives": "executives.xlsx",
    "board_members": "board_members.xlsx",
    "accountant": "accountant.xlsx",
    "dfo_status": "dfo_status.xlsx",
}
def append_rows_with_sheet_split(workbook, sheet_title: str, headers: Sequence[str], rows: Sequence[Sequence]):
    sheet_index = 1
    ws = workbook.active
    ws.title = sheet_title
    ws.append(list(headers))
    written_on_sheet = 0

    for row in rows:
        if written_on_sheet >= MAX_DATA_ROWS_PER_SHEET:
            sheet_index += 1
            ws = workbook.create_sheet(f"{sheet_title} {sheet_index}")
            ws.append(list(headers))
            written_on_sheet = 0

        ws.append(list(row))
        written_on_sheet += 1


def save_workbook(output_path: str, sheet_title: str, headers: Sequence[str], rows: Sequence[Sequence]):
    wb = openpyxl.Workbook()
    append_rows_with_sheet_split(wb, sheet_title, headers, rows)
    wb.save(output_path)
    print(f"Saved {output_path}: {len(rows)} rows")
def build_tax_exports():
    for output_name, sheet_title in TAX_SOURCE_FILES.items():
        paths = [output_name, NEW_TAX_SOURCE_FILES[output_name]]
        headers, rows = merge_excel_files(paths)
        if not headers:
            print(f"Skipping {output_name}: no source files")
            continue

        save_workbook(
            os.path.join(OUTPUT_DIR, output_name),
            sheet_title,
            headers,
            rows,
        )


def read_excel_rows(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def merge_excel_files(paths: Sequence[str]):
    headers = None
    merged_rows: List[Sequence] = []
    seen_rows = set()
    collected_rows = []

    for path in paths:
        if not os.path.exists(path):
            continue

        rows = read_excel_rows(path)
        if not rows:
            continue

        source_headers = list(rows[0])
        if headers is None:
            headers = list(source_headers)
        else:
            for header in source_headers:
                if header not in headers:
                    headers.append(header)

        collected_rows.append((source_headers, rows[1:]))

    if headers is None:
        return [], []

    for source_headers, source_rows in collected_rows:
        header_index = {header: idx for idx, header in enumerate(source_headers)}
        for row in source_rows:
            normalized_row = [
                row[header_index[header]] if header in header_index and header_index[header] < len(row) else ""
                for header in headers
            ]
            row_key = tuple(normalized_row)
            if row_key in seen_rows:
                continue
            seen_rows.add(row_key)
            merged_rows.append(normalized_row)

    return headers, merged_rows


def build_affiliation_exports():
    for base_name, output_name in AFFILIATION_OUTPUTS.items():
        paths = [
            os.path.join("affiliation_dfo_excels", f"{base_name}.xlsx"),
            os.path.join("affiliation_dfo_excels_new", f"{base_name}_new.xlsx"),
        ]
        headers, rows = merge_excel_files(paths)
        if not headers:
            print(f"Skipping {base_name}: no source files")
            continue

        save_workbook(
            os.path.join(OUTPUT_DIR, output_name),
            "data",
            headers,
            rows,
        )


def build_financial_export():
    headers, rows = merge_excel_files(FINANCIAL_FILES)
    if not headers:
        print("Skipping financial_reports export: no source files")
        return

    save_workbook(
        os.path.join(OUTPUT_DIR, "bin_financial.xlsx"),
        "BIN_URLS",
        headers,
        rows,
    )


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    build_tax_exports()
    build_affiliation_exports()
    build_financial_export()
    print(f"Done. Combined files are in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
