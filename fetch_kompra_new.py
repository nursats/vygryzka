import json
import os
import time
from typing import Iterable

import openpyxl
import requests

INPUT_DIR = "new"
API_TOKEN = "G1BCW0TvPlFFf7jM7wmLoi"
API_URL = "https://kompra.kz/api/v2/tax-details"
JSON_DIR = "json_responses_new"
PROGRESS_FILE = "progress_new.json"
MAX_RETRIES = 2
RETRY_DELAY = 2
REQUEST_DELAY = 0.5


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed": []}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def iter_excel_files() -> Iterable[str]:
    for name in sorted(os.listdir(INPUT_DIR)):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            yield os.path.join(INPUT_DIR, name)


def normalize_bin(value):
    if value is None:
        return None

    if isinstance(value, int):
        candidate = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            return None
        candidate = str(int(value))
    else:
        candidate = str(value).strip()

    if not candidate:
        return None

    if candidate.endswith(".0") and candidate[:-2].isdigit():
        candidate = candidate[:-2]

    candidate = "".join(ch for ch in candidate if ch.isdigit())
    if len(candidate) != 12:
        return None

    return candidate


def read_bins_from_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    bins = []
    seen = set()
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        bin_id = normalize_bin(row[0])
        if bin_id and bin_id not in seen:
            bins.append(bin_id)
            seen.add(bin_id)

    wb.close()
    return bins


def read_all_bins():
    bins = []
    seen = set()

    for path in iter_excel_files():
        file_bins = read_bins_from_file(path)
        print(f"{os.path.basename(path)}: found {len(file_bins)} BINs")
        for bin_id in file_bins:
            if bin_id not in seen:
                bins.append(bin_id)
                seen.add(bin_id)

    return bins


def fetch_tax_details(identifier):
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.get(
                API_URL,
                params={"identifier": identifier, "api-token": API_TOKEN},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()

            top_status = data.get("status")
            payments = data.get("content", {}).get("answer", {}).get("payment", [])

            if len(payments) == 0 and top_status is None:
                if attempt < MAX_RETRIES - 1:
                    print(
                        f"    Empty response (status=null) for {identifier}, "
                        f"retrying in {RETRY_DELAY}s (attempt {attempt + 1}/{MAX_RETRIES})..."
                    )
                    time.sleep(RETRY_DELAY)
                    continue

                print(f"    Max retries reached for {identifier}, skipping")
                return None

            if len(payments) == 0 and top_status is False:
                print(f"    No tax data for {identifier} (status=false), saving")

            return data

        except requests.exceptions.RequestException as e:
            if attempt < MAX_RETRIES - 1:
                print(f"    Request error for {identifier}: {e}, retrying in {RETRY_DELAY}s...")
                time.sleep(RETRY_DELAY)
            else:
                print(f"    Failed after {MAX_RETRIES} attempts for {identifier}: {e}")
                return None

    return None


def main():
    os.makedirs(JSON_DIR, exist_ok=True)

    progress = load_progress()
    completed = set(progress.get("completed", []))

    bins = read_all_bins()
    remaining = [bin_id for bin_id in bins if bin_id not in completed]

    print(f"Total unique BINs to process: {len(bins)}")
    print(f"Already completed: {len(completed)}")
    print(f"Remaining: {len(remaining)}")

    for index, bin_id in enumerate(remaining, start=1):
        print(f"[{len(completed) + index}/{len(bins)}] Fetching {bin_id}...")
        data = fetch_tax_details(bin_id)

        if data is None:
            print(f"    Skipped {bin_id} (no data), will retry next run")
            time.sleep(REQUEST_DELAY)
            continue

        json_path = os.path.join(JSON_DIR, f"{bin_id}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        progress.setdefault("completed", []).append(bin_id)
        completed.add(bin_id)
        save_progress(progress)
        time.sleep(REQUEST_DELAY)

    print(f"\nFinished. JSON files are in: {JSON_DIR}")


if __name__ == "__main__":
    main()
