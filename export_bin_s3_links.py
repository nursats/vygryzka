import argparse
import os
import re
from typing import Iterable, List, Sequence

import openpyxl
from sqlalchemy import bindparam, create_engine, text
from sqlalchemy.engine import Engine

DEFAULT_INPUT_EXCEL = "20260304 KOMPRA (2).xlsx"
DEFAULT_OUTPUT_EXCEL = "bin_s3_links.xlsx"
DEFAULT_INPUT_DIR = "new"
DEFAULT_SHEET_NAME = "Phase 1"
DEFAULT_START_ROW = 3
DEFAULT_COLUMN = 1
DEFAULT_CHUNK_SIZE = 1000


def normalize_bin(value) -> str:
    if value is None:
        return ""

    if isinstance(value, int):
        raw = str(value)
    elif isinstance(value, float):
        raw = str(int(value)) if value.is_integer() else str(value)
    else:
        raw = str(value).strip()

    raw = raw.replace("\u00a0", "").replace(" ", "")
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]

    if raw.isdigit() and len(raw) <= 12:
        return raw.zfill(12)

    return raw


def read_bins_from_excel(
    file_path: str,
    sheet_name: str,
    start_row: int,
    column: int,
) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"Sheet '{sheet_name}' not found. Using '{ws.title}' instead.")

    bins: List[str] = []
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row,
        min_col=column,
        max_col=column,
        values_only=True,
    ):
        bin_value = normalize_bin(row[0])
        if bin_value:
            bins.append(bin_value)

    wb.close()

    seen = set()
    unique_bins = []
    for b in bins:
        if b not in seen:
            unique_bins.append(b)
            seen.add(b)

    return unique_bins


def read_bins_from_dir(
    dir_path: str,
    sheet_name: str,
    start_row: int,
    column: int,
) -> List[str]:
    excel_paths = sorted(
        os.path.join(dir_path, name)
        for name in os.listdir(dir_path)
        if name.lower().endswith(".xlsx") and not name.startswith("~$")
    )
    if not excel_paths:
        raise FileNotFoundError(f"No .xlsx files found in directory: {dir_path}")

    bins: List[str] = []
    seen = set()

    for excel_path in excel_paths:
        file_bins = read_bins_from_excel(
            file_path=excel_path,
            sheet_name=sheet_name,
            start_row=start_row,
            column=column,
        )
        print(f"{os.path.basename(excel_path)}: {len(file_bins)} unique BINs")
        for bin_value in file_bins:
            if bin_value not in seen:
                bins.append(bin_value)
                seen.add(bin_value)

    return bins


def get_db_settings() -> dict:
    return {
        "name": os.getenv("DB_NAME", "kompra"),
        "user": os.getenv("DB_USER", "python"),
        "password": os.getenv("DB_PASS", "python123"),
        "host": os.getenv("DB_HOST", "10.0.0.1"),
        "port": os.getenv("DB_PORT", "5432"),
        "schema": os.getenv("DB_SCHEMA", "public"),
    }


def create_db_engine(settings: dict) -> Engine:
    url = "postgresql://{}:{}@{}:{}/{}".format(
        settings["user"],
        settings["password"],
        settings["host"],
        settings["port"],
        settings["name"],
    )
    return create_engine(url, client_encoding="utf8")


def chunked(values: Sequence[str], chunk_size: int) -> Iterable[Sequence[str]]:
    for i in range(0, len(values), chunk_size):
        yield values[i : i + chunk_size]


def validate_schema(schema: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", schema):
        raise ValueError(f"Invalid schema name: {schema!r}")
    return schema


def fetch_bin_urls(
    engine: Engine,
    bins: Sequence[str],
    schema: str,
    chunk_size: int,
) -> dict:
    """Return {bin: url} with one URL per BIN (latest report_date)."""
    if not bins:
        return {}

    schema = validate_schema(schema)
    # Pick the single most-recent URL per BIN.
    stmt = text(
        f"""
        SELECT DISTINCT ON (identifier) identifier AS bin, doc_url AS url
        FROM {schema}.financial_reports
        WHERE identifier IN :bins
          AND doc_url IS NOT NULL
          AND btrim(doc_url) <> ''
        ORDER BY identifier, report_date DESC
        """
    ).bindparams(bindparam("bins", expanding=True))

    bin_url: dict = {}
    with engine.connect() as conn:
        for bin_chunk in chunked(bins, chunk_size):
            rows = conn.execute(stmt, {"bins": list(bin_chunk)}).mappings().all()
            for row in rows:
                bin_url[row["bin"]] = row["url"]

    return bin_url


def fetch_all_bin_urls(
    engine: Engine,
    bins: Sequence[str],
    schema: str,
    chunk_size: int,
) -> List[dict]:
    """Return all URL rows for all BINs."""
    if not bins:
        return []

    schema = validate_schema(schema)
    stmt = text(
        f"""
        SELECT identifier AS bin, doc_url AS url, report_date
        FROM {schema}.financial_reports
        WHERE identifier IN :bins
          AND doc_url IS NOT NULL
          AND btrim(doc_url) <> ''
        ORDER BY identifier, report_date DESC, id DESC
        """
    ).bindparams(bindparam("bins", expanding=True))

    result_rows: List[dict] = []
    with engine.connect() as conn:
        for bin_chunk in chunked(bins, chunk_size):
            rows = conn.execute(stmt, {"bins": list(bin_chunk)}).mappings().all()
            result_rows.extend(dict(row) for row in rows)

    return result_rows


def write_output_excel(bins: Sequence[str], bin_url: dict, output_path: str) -> None:
    """Write output preserving input Excel BIN order, one row per BIN."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BIN_URLS"

    ws.append(["bin", "url"])
    for bin_value in bins:
        url = bin_url.get(bin_value)
        if url:
            ws.append([bin_value, url])

    wb.save(output_path)


def write_output_excel_all_rows(rows: Sequence[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BIN_URLS"

    ws.append(["bin", "url", "report_date"])
    for row in rows:
        ws.append([row["bin"], row["url"], row.get("report_date")])

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export BIN + S3 URL pairs from DB for BINs listed in Excel."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT_EXCEL, help="Path to input Excel with BINs")
    parser.add_argument(
        "--input-dir",
        default=None,
        help="Directory with input Excels containing BINs in the selected column",
    )
    parser.add_argument("--output", default=DEFAULT_OUTPUT_EXCEL, help="Path to output Excel")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Sheet name with BINs")
    parser.add_argument("--start-row", type=int, default=DEFAULT_START_ROW, help="Data start row")
    parser.add_argument("--column", type=int, default=DEFAULT_COLUMN, help="BIN column number (1-based)")
    parser.add_argument("--schema", default=None, help="DB schema (default: DB_SCHEMA env or public)")
    parser.add_argument("--chunk-size", type=int, default=DEFAULT_CHUNK_SIZE, help="BIN query chunk size")
    parser.add_argument(
        "--all-links",
        action="store_true",
        help="Export all rows from financial_reports instead of only the latest URL per BIN",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    settings = get_db_settings()
    if args.schema:
        settings["schema"] = args.schema

    if args.input_dir:
        if not os.path.isdir(args.input_dir):
            raise FileNotFoundError(f"Input directory not found: {args.input_dir}")
        bins = read_bins_from_dir(
            dir_path=args.input_dir,
            sheet_name=args.sheet,
            start_row=args.start_row,
            column=args.column,
        )
        print(f"Read {len(bins)} unique BINs from directory")
    else:
        if not os.path.exists(args.input):
            raise FileNotFoundError(f"Input file not found: {args.input}")

        bins = read_bins_from_excel(
            file_path=args.input,
            sheet_name=args.sheet,
            start_row=args.start_row,
            column=args.column,
        )

        print(f"Read {len(bins)} unique BINs from Excel")

    engine = create_db_engine(settings)
    if args.all_links:
        rows = fetch_all_bin_urls(
            engine=engine,
            bins=bins,
            schema=settings["schema"],
            chunk_size=max(1, args.chunk_size),
        )
        write_output_excel_all_rows(rows, args.output)
        found_bins = len({row["bin"] for row in rows})
        skipped = len(bins) - found_bins
        print(f"Rows exported: {len(rows)}")
        print(f"BINs with links: {found_bins}")
        print(f"BINs skipped (no link): {skipped}")
    else:
        bin_url = fetch_bin_urls(
            engine=engine,
            bins=bins,
            schema=settings["schema"],
            chunk_size=max(1, args.chunk_size),
        )
        write_output_excel(bins, bin_url, args.output)
        found = len(bin_url)
        skipped = len(bins) - found
        print(f"Rows exported: {found}")
        print(f"BINs with links: {found}")
        print(f"BINs skipped (no link): {skipped}")

    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
